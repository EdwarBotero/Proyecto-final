"""
Sistema de Parqueadero Mejorado - Módulo de Base de Datos
=========================================================

Este módulo gestiona toda la lógica de negocio y persistencia de datos del sistema de parqueadero.
Implementa las operaciones CRUD (Crear, Leer, Actualizar, Eliminar) para los vehículos y el historial,
así como los cálculos de tarifas, duración y valores a pagar.

"""

# Importación de módulos necesarios
import sqlite3  # Módulo para trabajar con bases de datos SQLite
import csv  # Módulo para exportar datos en formato CSV
from pathlib import Path  # Módulo para manejar rutas de archivos de forma segura
from datetime import datetime, date, time  # Módulos para manejar fechas y horas

# Definición de constantes
# Ruta donde se almacena la base de datos, utilizando Path para compatibilidad multiplataforma
DB_PATH = Path("data/parking.db")

def conectar():
    """
    Establece conexión con la base de datos y crea las tablas si no existen.
    
    Esta función es fundamental para el sistema ya que:
    1. Asegura que el directorio de datos exista
    2. Establece la conexión con la base de datos SQLite
    3. Crea las tablas necesarias si no existen
    4. Inicializa las tarifas por defecto si es necesario
    
    Returns:
        sqlite3.Connection: Objeto de conexión a la base de datos
    """
    # Creamos el directorio padre si no existe
    DB_PATH.parent.mkdir(exist_ok=True)
    
    # Establecemos conexión con la base de datos
    conn = sqlite3.connect(DB_PATH)
    
    # Creamos la tabla de vehículos activos con soporte para fecha y hora completas
    # Esta tabla almacena los vehículos que actualmente están en el parqueadero
    conn.execute('''
        CREATE TABLE IF NOT EXISTS vehiculos (
            placa TEXT PRIMARY KEY,  -- Placa del vehículo como clave primaria
            fecha_entrada TEXT,      -- Fecha de entrada en formato YYYY-MM-DD
            hora_entrada INTEGER,    -- Hora de entrada (0-23)
            minuto_entrada INTEGER,  -- Minuto de entrada (0-59)
            tipo TEXT CHECK (tipo IN ('Carro', 'Moto')),  -- Tipo de vehículo con restricción
            usuario TEXT             -- Usuario que registró la entrada
        )
    ''')
    
    # Creamos la tabla de historial con soporte para fecha y hora completas
    # Esta tabla almacena el registro histórico de todos los vehículos que han usado el parqueadero
    conn.execute('''
        CREATE TABLE IF NOT EXISTS historial (
            id INTEGER PRIMARY KEY AUTOINCREMENT,  -- ID único autoincremental
            placa TEXT,                -- Placa del vehículo
            tipo TEXT,                 -- Tipo de vehículo
            fecha_entrada TEXT,        -- Fecha de entrada en formato YYYY-MM-DD
            hora_entrada INTEGER,      -- Hora de entrada (0-23)
            minuto_entrada INTEGER,    -- Minuto de entrada (0-59)
            fecha_salida TEXT,         -- Fecha de salida en formato YYYY-MM-DD
            hora_salida INTEGER,       -- Hora de salida (0-23)
            minuto_salida INTEGER,     -- Minuto de salida (0-59)
            duracion_horas REAL,       -- Duración en horas (con decimales para fracciones)
            valor_pagado REAL,         -- Valor pagado
            usuario_registro TEXT,     -- Usuario que registró la salida
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP  -- Fecha y hora del registro
        )
    ''')
    
    # Creamos tabla de tarifas
    # Esta tabla permite configurar diferentes tarifas según tipo de vehículo, día y hora
    conn.execute('''
        CREATE TABLE IF NOT EXISTS tarifas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,  -- ID único autoincremental
            tipo_vehiculo TEXT,        -- Tipo de vehículo (Carro o Moto)
            dia_semana TEXT,           -- Día de la semana o 'todos'
            hora_inicio INTEGER,       -- Hora de inicio para esta tarifa
            hora_fin INTEGER,          -- Hora de fin para esta tarifa
            tarifa_hora REAL,          -- Valor por hora completa
            tarifa_fraccion REAL,      -- Valor por fracción (15 minutos)
            activo INTEGER DEFAULT 1   -- Indica si la tarifa está activa (1) o no (0)
        )
    ''')
    
    # Insertamos tarifas por defecto si no existen
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM tarifas")
    if cursor.fetchone()[0] == 0:
        # Tarifa para carros (todos los días)
        conn.execute(
            "INSERT INTO tarifas (tipo_vehiculo, dia_semana, hora_inicio, hora_fin, tarifa_hora, tarifa_fraccion) VALUES (?, ?, ?, ?, ?, ?)",
            ("Carro", "todos", 0, 23, 5000, 1200)  # $5000 por hora, $1200 por fracción
        )
        # Tarifa para motos (todos los días)
        conn.execute(
            "INSERT INTO tarifas (tipo_vehiculo, dia_semana, hora_inicio, hora_fin, tarifa_hora, tarifa_fraccion) VALUES (?, ?, ?, ?, ?, ?)",
            ("Moto", "todos", 0, 23, 3500, 900)  # $3500 por hora, $900 por fracción
        )
    
    # Guardamos los cambios y retornamos la conexión
    conn.commit()
    return conn

def obtener_fecha_hora_actual():
    """
    Retorna la fecha y hora actual en formato adecuado para la BD.
    
    Esta función es utilizada para obtener la fecha y hora actual del sistema
    en el formato requerido por la base de datos.
    
    Returns:
        tuple: Tupla con (fecha_actual, hora_actual, minuto_actual)
            - fecha_actual (str): Fecha en formato 'YYYY-MM-DD'
            - hora_actual (int): Hora en formato 24h (0-23)
            - minuto_actual (int): Minutos (0-59)
    """
    # Obtenemos la fecha y hora actual
    ahora = datetime.now()
    
    # Formateamos la fecha como YYYY-MM-DD
    fecha_actual = ahora.strftime("%Y-%m-%d")
    
    # Extraemos la hora y minuto
    hora_actual = ahora.hour
    minuto_actual = ahora.minute
    
    return fecha_actual, hora_actual, minuto_actual

def calcular_duracion(fecha_entrada, hora_entrada, minuto_entrada, fecha_salida, hora_salida, minuto_salida):
    """
    Calcula la duración en horas (con decimales para fracciones) entre entrada y salida.
    
    Esta función calcula con precisión el tiempo que un vehículo ha permanecido en el
    parqueadero, considerando fechas diferentes y expresando el resultado en horas con
    decimales para representar fracciones de hora.
    
    Args:
        fecha_entrada (str): Fecha de entrada en formato 'YYYY-MM-DD'
        hora_entrada (int): Hora de entrada (0-23)
        minuto_entrada (int): Minuto de entrada (0-59)
        fecha_salida (str): Fecha de salida en formato 'YYYY-MM-DD'
        hora_salida (int): Hora de salida (0-23)
        minuto_salida (int): Minuto de salida (0-59)
    
    Returns:
        float: Duración en horas con decimales (ej: 2.5 para 2 horas y 30 minutos)
    """
    # Convertimos las cadenas de fecha a objetos datetime
    entrada = datetime.combine(
        datetime.strptime(fecha_entrada, "%Y-%m-%d").date(),
        time(hour=hora_entrada, minute=minuto_entrada)
    )
    
    salida = datetime.combine(
        datetime.strptime(fecha_salida, "%Y-%m-%d").date(),
        time(hour=hora_salida, minute=minuto_salida)
    )
    
    # Si la salida es anterior a la entrada, asumimos que pasó al menos un día
    # Esto maneja el caso donde un vehículo entra un día y sale al siguiente
    if salida < entrada:
        salida = salida.replace(day=salida.day + 1)
    
    # Calculamos la diferencia en horas (con decimales)
    # total_seconds() da la diferencia en segundos, dividimos por 3600 para obtener horas
    diferencia = salida - entrada
    horas = diferencia.total_seconds() / 3600
    
    # Redondeamos a 2 decimales para mejor legibilidad
    return round(horas, 2)

def obtener_tarifa(tipo_vehiculo, fecha, hora):
    """
    Obtiene la tarifa aplicable según tipo de vehículo, día y hora.
    
    Esta función implementa un sistema flexible de tarifas que permite
    configurar diferentes precios según:
    - Tipo de vehículo (Carro o Moto)
    - Día de la semana
    - Hora del día
    
    Args:
        tipo_vehiculo (str): Tipo de vehículo ('Carro' o 'Moto')
        fecha (str): Fecha en formato 'YYYY-MM-DD'
        hora (int): Hora en formato 24h (0-23)
    
    Returns:
        tuple: Tupla con (tarifa_hora, tarifa_fraccion)
            - tarifa_hora (float): Tarifa por hora completa
            - tarifa_fraccion (float): Tarifa por fracción (15 minutos)
    """
    try:
        # Establecemos conexión con la base de datos
        conn = conectar()
        cursor = conn.cursor()
        
        # Obtenemos el día de la semana (0=lunes, 6=domingo)
        dia_semana = datetime.strptime(fecha, "%Y-%m-%d").weekday()
        dias = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]
        
        # Buscamos tarifa específica para este día y hora
        # La consulta prioriza tarifas específicas para el día sobre tarifas genéricas ('todos')
        cursor.execute("""
            SELECT tarifa_hora, tarifa_fraccion FROM tarifas 
            WHERE tipo_vehiculo = ? AND 
                  (dia_semana = ? OR dia_semana = 'todos') AND
                  hora_inicio <= ? AND hora_fin >= ? AND
                  activo = 1
            ORDER BY CASE WHEN dia_semana = ? THEN 0 ELSE 1 END
            LIMIT 1
        """, (tipo_vehiculo, dias[dia_semana], hora, hora, dias[dia_semana]))
        
        resultado = cursor.fetchone()
        if resultado:
            return resultado
        
        # Si no hay tarifa específica, usamos la tarifa por defecto
        # Carros: $5000/hora, $1200/fracción
        # Motos: $3500/hora, $900/fracción
        return (5000, 1200) if tipo_vehiculo == "Carro" else (3500, 900)
    
    except sqlite3.Error as e:
        # En caso de error, registramos el problema y retornamos tarifas por defecto
        print(f"Error al obtener tarifa: {e}")
        return (5000, 1200) if tipo_vehiculo == "Carro" else (3500, 900)

def calcular_valor(tipo_vehiculo, duracion_horas, fecha_entrada, hora_entrada):
    """
    Calcula el valor a pagar según tipo de vehículo y duración.
    
    Esta función implementa un algoritmo de cálculo de tarifas que considera:
    - Horas completas
    - Fracciones de hora (en intervalos de 15 minutos)
    - Tarifas diferenciadas por tipo de vehículo
    
    Args:
        tipo_vehiculo (str): Tipo de vehículo ('Carro' o 'Moto')
        duracion_horas (float): Duración en horas con decimales
        fecha_entrada (str): Fecha de entrada en formato 'YYYY-MM-DD'
        hora_entrada (int): Hora de entrada (0-23)
    
    Returns:
        float: Valor a pagar
    """
    try:
        # Obtenemos la tarifa aplicable para este vehículo, fecha y hora
        tarifa_hora, tarifa_fraccion = obtener_tarifa(tipo_vehiculo, fecha_entrada, hora_entrada)
        
        # Calculamos horas completas y fracción
        horas_completas = int(duracion_horas)  # Parte entera (horas completas)
        fraccion = duracion_horas - horas_completas  # Parte decimal (fracción de hora)
        
        # Si hay fracción, aplicamos tarifa por fracción (cada 15 minutos)
        # El sistema cobra por intervalos de 15 minutos (0.25 horas)
        if fraccion > 0:
            if fraccion <= 0.25:  # Hasta 15 minutos
                valor_fraccion = tarifa_fraccion
            elif fraccion <= 0.5:  # Hasta 30 minutos
                valor_fraccion = tarifa_fraccion * 2
            elif fraccion <= 0.75:  # Hasta 45 minutos
                valor_fraccion = tarifa_fraccion * 3
            else:  # Más de 45 minutos
                valor_fraccion = tarifa_hora  # Cobra hora completa
        else:
            valor_fraccion = 0
        
        # Calculamos valor total: (horas completas * tarifa por hora) + valor por fracción
        valor_total = (horas_completas * tarifa_hora) + valor_fraccion
        
        # Redondeamos para evitar decimales en el valor final
        return round(valor_total, 0)
    
    except Exception as e:
        # En caso de error, registramos el problema y usamos un cálculo simplificado
        print(f"Error en cálculo de valor: {e}")
        # Valores por defecto en caso de error
        tarifas = {"Carro": 5000, "Moto": 3500}
        return tarifas.get(tipo_vehiculo, 3000) * duracion_horas

def validar_placa(placa):
    """
    Valida el formato de la placa según normativa colombiana.
    
    Esta función implementa una validación básica del formato de placas:
    - Longitud entre 5 y 7 caracteres
    
    Args:
        placa (str): Placa del vehículo a validar
    
    Returns:
        bool: True si la placa es válida, False en caso contrario
    """
    # Verificamos que la placa no esté vacía y tenga la longitud adecuada
    if not placa or len(placa) < 5 or len(placa) > 7:
        return False
    
    # Formato básico: 3 letras seguidas de 3 números (carros) o 2 letras, 2 números y 1 letra (motos)
    # Nota: En esta versión simplificada, solo verificamos la longitud
    return True

def registrar_ingreso(placa, fecha=None, hora=None, minuto=None, tipo="Carro", usuario="sistema"):
    """
    Registra el ingreso de un vehículo al parqueadero.
    
    Esta función realiza las siguientes operaciones:
    1. Valida el formato de la placa
    2. Verifica que el vehículo no esté ya en el parqueadero
    3. Registra la entrada con fecha y hora
    
    Args:
        placa (str): Placa del vehículo
        fecha (str, optional): Fecha de entrada en formato 'YYYY-MM-DD'. Si es None, usa la fecha actual.
        hora (int, optional): Hora de entrada (0-23). Si es None, usa la hora actual.
        minuto (int, optional): Minuto de entrada (0-59). Si es None, usa el minuto actual.
        tipo (str, optional): Tipo de vehículo ('Carro' o 'Moto'). Por defecto 'Carro'.
        usuario (str, optional): Usuario que registra la entrada. Por defecto 'sistema'.
    
    Returns:
        tuple: (success, message)
            - success (bool): True si la operación fue exitosa, False en caso contrario
            - message (str): Mensaje descriptivo del resultado
    """
    # Validamos el formato de la placa
    if not validar_placa(placa):
        return False, "Formato de placa inválido"
    
    try:
        # Establecemos conexión con la base de datos
        conn = conectar()
        cursor = conn.cursor()
        
        # Verificar si la placa ya está registrada
        cursor.execute("SELECT placa FROM vehiculos WHERE placa=?", (placa,))
        if cursor.fetchone():
            return False, "La placa ya está registrada en el parqueadero"
        
        # Si no se proporcionan fecha y hora, usamos la actual
        if fecha is None or hora is None or minuto is None:
            fecha, hora, minuto = obtener_fecha_hora_actual()
        
        # Validamos el tipo de vehículo
        tipo = tipo.capitalize()
        if tipo not in ["Carro", "Moto"]:
            return False, "Tipo de vehículo inválido"
        
        # Insertar el vehículo en la tabla de vehículos activos
        conn.execute(
            "INSERT INTO vehiculos (placa, fecha_entrada, hora_entrada, minuto_entrada, tipo, usuario) VALUES (?, ?, ?, ?, ?, ?)",
            (placa, fecha, hora, minuto, tipo, usuario)
        )
        conn.commit()
        return True, "Ingreso registrado correctamente"
    
    except sqlite3.Error as e:
        # En caso de error en la base de datos, registramos el problema
        print(f"Error en registrar ingreso: {e}")
        return False, f"Error en la base de datos: {e}"

def registrar_salida(placa, fecha_salida=None, hora_salida=None, minuto_salida=None, usuario="sistema"):
    """
    Registra la salida de un vehículo del parqueadero.
    
    Esta función realiza las siguientes operaciones:
    1. Verifica que el vehículo esté en el parqueadero
    2. Calcula la duración de la estancia y el valor a pagar
    3. Registra la salida en el historial
    4. Elimina el vehículo de la lista de activos
    
    Args:
        placa (str): Placa del vehículo
        fecha_salida (str, optional): Fecha de salida en formato 'YYYY-MM-DD'. Si es None, usa la fecha actual.
        hora_salida (int, optional): Hora de salida (0-23). Si es None, usa la hora actual.
        minuto_salida (int, optional): Minuto de salida (0-59). Si es None, usa el minuto actual.
        usuario (str, optional): Usuario que registra la salida. Por defecto 'sistema'.
    
    Returns:
        tuple: (success, result)
            - success (bool): True si la operación fue exitosa, False en caso contrario
            - result: Si success es True, un diccionario con información de la salida.
                      Si success es False, un mensaje de error.
    """
    try:
        # Establecemos conexión con la base de datos
        conn = conectar()
        cursor = conn.cursor()
        
        # Verificar si el vehículo está en el parqueadero
        cursor.execute(
            "SELECT fecha_entrada, hora_entrada, minuto_entrada, tipo FROM vehiculos WHERE placa=?",
            (placa,)
        )
        fila = cursor.fetchone()
        if not fila:
            return False, "Vehículo no encontrado en el parqueadero"
        
        # Extraemos los datos de entrada
        fecha_entrada, hora_entrada, minuto_entrada, tipo = fila
        
        # Si no se proporcionan fecha y hora de salida, usamos la actual
        if fecha_salida is None or hora_salida is None or minuto_salida is None:
            fecha_salida, hora_salida, minuto_salida = obtener_fecha_hora_actual()
        
        # Calculamos la duración y el valor a pagar
        duracion = calcular_duracion(
            fecha_entrada, hora_entrada, minuto_entrada,
            fecha_salida, hora_salida, minuto_salida
        )
        
        valor = calcular_valor(tipo, duracion, fecha_entrada, hora_entrada)
        
        # Registramos en el historial
        cursor.execute(
            """
            INSERT INTO historial 
            (placa, tipo, fecha_entrada, hora_entrada, minuto_entrada, 
             fecha_salida, hora_salida, minuto_salida, duracion_horas, valor_pagado, usuario_registro)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (placa, tipo, fecha_entrada, hora_entrada, minuto_entrada,
             fecha_salida, hora_salida, minuto_salida, duracion, valor, usuario)
        )
        
        # Eliminamos el vehículo de la tabla de vehículos activos
        cursor.execute("DELETE FROM vehiculos WHERE placa=?", (placa,))
        conn.commit()
        
        # Retornamos éxito y datos relevantes
        return True, {
            "duracion": duracion,
            "valor": valor,
            "tipo": tipo,
            "entrada": f"{fecha_entrada} {hora_entrada}:{minuto_entrada:02d}",
            "salida": f"{fecha_salida} {hora_salida}:{minuto_salida:02d}"
        }
    
    except sqlite3.Error as e:
        # En caso de error en la base de datos, registramos el problema
        print(f"Error en registrar salida: {e}")
        return False, f"Error en la base de datos: {e}"

def obtener_vehiculos_activos():
    """
    Obtiene la lista de vehículos actualmente en el parqueadero.
    
    Returns:
        list: Lista de tuplas con información de vehículos activos.
              Cada tupla contiene (placa, tipo, fecha_entrada, hora_entrada, minuto_entrada)
    """
    try:
        # Establecemos conexión con la base de datos
        conn = conectar()
        cursor = conn.cursor()
        
        # Consultamos todos los vehículos activos ordenados por tiempo de entrada
        cursor.execute("""
            SELECT placa, tipo, fecha_entrada, hora_entrada, minuto_entrada
            FROM vehiculos
            ORDER BY fecha_entrada, hora_entrada, minuto_entrada
        """)
        return cursor.fetchall()
    except sqlite3.Error as e:
        # En caso de error, registramos el problema y retornamos lista vacía
        print(f"Error al obtener vehículos activos: {e}")
        return []

def obtener_historial(filtro_placa=None, filtro_tipo=None, fecha_inicio=None, fecha_fin=None):
    """
    Obtiene el historial de vehículos con opciones de filtrado.
    
    Esta función permite filtrar el historial por:
    - Placa (parcial o completa)
    - Tipo de vehículo
    - Rango de fechas
    
    Args:
        filtro_placa (str, optional): Filtro por placa (parcial o completa)
        filtro_tipo (str, optional): Filtro por tipo ('Carro' o 'Moto')
        fecha_inicio (str, optional): Fecha inicial para filtrar (formato 'YYYY-MM-DD')
        fecha_fin (str, optional): Fecha final para filtrar (formato 'YYYY-MM-DD')
    
    Returns:
        list: Lista de tuplas con información del historial filtrado
    """
    try:
        # Establecemos conexión con la base de datos
        conn = conectar()
        cursor = conn.cursor()
        
        # Construimos la consulta base
        query = """
            SELECT placa, tipo, fecha_entrada, hora_entrada, minuto_entrada,
                   fecha_salida, hora_salida, minuto_salida, duracion_horas, valor_pagado
            FROM historial
            WHERE 1=1
        """
        params = []
        
        # Añadimos filtros si se proporcionan
        if filtro_placa:
            query += " AND placa LIKE ?"
            params.append(f"%{filtro_placa}%")
        
        if filtro_tipo:
            query += " AND tipo = ?"
            params.append(filtro_tipo)
        
        if fecha_inicio:
            query += " AND fecha_entrada >= ?"
            params.append(fecha_inicio)
        
        if fecha_fin:
            query += " AND fecha_entrada <= ?"
            params.append(fecha_fin)
        
        # Ordenamos por fecha de registro descendente (más recientes primero)
        query += " ORDER BY fecha_registro DESC"
        
        cursor.execute(query, params)
        return cursor.fetchall()
    
    except sqlite3.Error as e:
        # En caso de error, registramos el problema y retornamos lista vacía
        print(f"Error en obtener historial: {e}")
        return []

def exportar_historial_csv(ruta="historial_parqueadero.csv", filtro_placa=None, filtro_tipo=None, fecha_inicio=None, fecha_fin=None):
    """
    Exporta el historial a un archivo CSV con opciones de filtrado.
    
    Esta función permite exportar el historial filtrado a un archivo CSV,
    ideal para análisis en hojas de cálculo o importación a otros sistemas.
    
    Args:
        ruta (str, optional): Ruta del archivo CSV a generar
        filtro_placa (str, optional): Filtro por placa (parcial o completa)
        filtro_tipo (str, optional): Filtro por tipo ('Carro' o 'Moto')
        fecha_inicio (str, optional): Fecha inicial para filtrar (formato 'YYYY-MM-DD')
        fecha_fin (str, optional): Fecha final para filtrar (formato 'YYYY-MM-DD')
    
    Returns:
        tuple: (success, message)
            - success (bool): True si la operación fue exitosa, False en caso contrario
            - message (str): Mensaje descriptivo del resultado
    """
    try:
        # Obtenemos el historial filtrado
        historial = obtener_historial(filtro_placa, filtro_tipo, fecha_inicio, fecha_fin)
        if not historial:
            return False, "No hay datos para exportar"
        
        # Creamos el archivo CSV
        with open(ruta, mode="w", newline="", encoding="utf-8") as archivo:
            writer = csv.writer(archivo)
            
            # Escribimos la cabecera
            writer.writerow([
                "Placa", "Tipo", "Fecha Entrada", "Hora Entrada", 
                "Fecha Salida", "Hora Salida", "Duración (horas)", "Valor"
            ])
            
            # Escribimos los datos
            for fila in historial:
                placa, tipo, fecha_entrada, hora_entrada, minuto_entrada, \
                fecha_salida, hora_salida, minuto_salida, duracion, valor = fila
                
                # Formateamos las horas para mejor legibilidad
                entrada_str = f"{fecha_entrada} {hora_entrada}:{minuto_entrada:02d}"
                salida_str = f"{fecha_salida} {hora_salida}:{minuto_salida:02d}" if fecha_salida else "N/A"
                
                writer.writerow([
                    placa, tipo, fecha_entrada, f"{hora_entrada}:{minuto_entrada:02d}",
                    fecha_salida, f"{hora_salida}:{minuto_salida:02d}" if fecha_salida else "N/A",
                    duracion, valor
                ])
        
        return True, f"Historial exportado a '{ruta}'"
    
    except Exception as e:
        # En caso de error, registramos el problema
        print(f"Error al exportar CSV: {e}")
        return False, f"Error al exportar: {e}"

def exportar_historial_excel(ruta="historial_parqueadero.xlsx", filtro_placa=None, filtro_tipo=None, fecha_inicio=None, fecha_fin=None):
    """
    Exporta el historial a un archivo Excel con opciones de filtrado.
    
    Esta función permite exportar el historial filtrado a un archivo Excel,
    con formato mejorado y estilos para mejor presentación.
    
    Args:
        ruta (str, optional): Ruta del archivo Excel a generar
        filtro_placa (str, optional): Filtro por placa (parcial o completa)
        filtro_tipo (str, optional): Filtro por tipo ('Carro' o 'Moto')
        fecha_inicio (str, optional): Fecha inicial para filtrar (formato 'YYYY-MM-DD')
        fecha_fin (str, optional): Fecha final para filtrar (formato 'YYYY-MM-DD')
    
    Returns:
        tuple: (success, message)
            - success (bool): True si la operación fue exitosa, False en caso contrario
            - message (str): Mensaje descriptivo del resultado
    """
    try:
        # Importamos openpyxl solo cuando se necesita
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Obtenemos el historial filtrado
        historial = obtener_historial(filtro_placa, filtro_tipo, fecha_inicio, fecha_fin)
        if not historial:
            return False, "No hay datos para exportar"
        
        # Creamos un nuevo libro y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Parqueadero"
        
        # Añadimos encabezados
        headers = [
            "Placa", "Tipo", "Fecha Entrada", "Hora Entrada", 
            "Fecha Salida", "Hora Salida", "Duración (horas)", "Valor"
        ]
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        centered = Alignment(horizontal="center")
        
        # Aplicamos encabezados con estilo
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
        
        # Añadimos datos
        for row_num, fila in enumerate(historial, 2):
            placa, tipo, fecha_entrada, hora_entrada, minuto_entrada, \
            fecha_salida, hora_salida, minuto_salida, duracion, valor = fila
            
            ws.cell(row=row_num, column=1, value=placa)
            ws.cell(row=row_num, column=2, value=tipo)
            ws.cell(row=row_num, column=3, value=fecha_entrada)
            ws.cell(row=row_num, column=4, value=f"{hora_entrada}:{minuto_entrada:02d}")
            ws.cell(row=row_num, column=5, value=fecha_salida if fecha_salida else "N/A")
            ws.cell(row=row_num, column=6, value=f"{hora_salida}:{minuto_salida:02d}" if fecha_salida else "N/A")
            ws.cell(row=row_num, column=7, value=duracion)
            
            # Formato para el valor
            cell = ws.cell(row=row_num, column=8, value=valor)
            cell.number_format = "$#,##0"
            cell.alignment = Alignment(horizontal="right")
        
        # Ajustamos anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardamos el archivo
        wb.save(ruta)
        return True, f"Historial exportado a '{ruta}'"
    
    except Exception as e:
        # En caso de error, registramos el problema
        print(f"Error al exportar Excel: {e}")
        return False, f"Error al exportar: {e}"

# Función para migración de datos antiguos (si es necesario)
def migrar_datos_antiguos():
    """
    Migra datos del formato antiguo al nuevo esquema.
    
    Esta función es útil cuando se actualiza el sistema desde una versión
    anterior que no tenía soporte para fechas completas y minutos.
    
    Returns:
        bool: True si la migración fue exitosa o no era necesaria, False en caso de error
    """
    try:
        # Establecemos conexión con la base de datos
        conn = conectar()
        cursor = conn.cursor()
        
        # Verificamos si existe la tabla historial antigua (sin campos de fecha)
        cursor.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='historial' AND 
            sql NOT LIKE '%fecha_entrada%'
        """)
        
        if cursor.fetchone():
            # La tabla existe en formato antiguo, necesitamos migrar
            print("Migrando datos antiguos...")
            
            # Obtenemos los datos antiguos
            cursor.execute("""
                SELECT placa, tipo, hora_entrada, hora_salida, duracion, valor_pagado
                FROM historial
            """)
            registros_antiguos = cursor.fetchall()
            
            # Renombramos la tabla antigua
            conn.execute("ALTER TABLE historial RENAME TO historial_old")
            
            # Creamos la nueva tabla con el esquema actualizado
            conn.execute('''
                CREATE TABLE historial (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    placa TEXT,
                    tipo TEXT,
                    fecha_entrada TEXT,
                    hora_entrada INTEGER,
                    minuto_entrada INTEGER,
                    fecha_salida TEXT,
                    hora_salida INTEGER,
                    minuto_salida INTEGER,
                    duracion_horas REAL,
                    valor_pagado REAL,
                    usuario_registro TEXT,
                    fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Fecha actual para los registros migrados
            fecha_actual = date.today().strftime("%Y-%m-%d")
            
            # Insertamos los datos antiguos en la nueva tabla
            for placa, tipo, hora_entrada, hora_salida, duracion, valor_pagado in registros_antiguos:
                conn.execute(
                    """
                    INSERT INTO historial 
                    (placa, tipo, fecha_entrada, hora_entrada, minuto_entrada, 
                     fecha_salida, hora_salida, minuto_salida, duracion_horas, valor_pagado, usuario_registro)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (placa, tipo, fecha_actual, hora_entrada, 0, 
                     fecha_actual, hora_salida, 0, duracion, valor_pagado, "migración")
                )
            
            conn.commit()
            print(f"Migración completada: {len(registros_antiguos)} registros")
            return True
        
        return False  # No era necesario migrar
    
    except sqlite3.Error as e:
        # En caso de error, registramos el problema
        print(f"Error en migración: {e}")
        return False


